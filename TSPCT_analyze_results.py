# ==========================================================
#  TSPCT_analyze_results.py
#
#  Statisticka analyza vysledkov z TSPCT_experiment_runner.py.
#  Cita 'Raw Results' z results/TSPCT_results.xlsx a DOPLNA
#  do toho isteho zosita dalsie harky:
#
#   - Friedman Test        (Friedman test pre kazde prostredie x nastavenie)
#   - Wilcoxon Posthoc      (parove porovnania s Holm korekciou)
#   - Convergence Charts    (priemerna konvergencia + 95% CI, vlozene ako obrazky)
#   - About                 (metodika, na co sa hodi ktory harok)
#
#  POUZITIE (az po TSPCT_experiment_runner.py):
#   python TSPCT_analyze_results.py
#
#  Vyzaduje: openpyxl, scipy, matplotlib
#   (pip install openpyxl scipy matplotlib --break-system-packages)
# ==========================================================

import json
import os
import math
import tempfile
from collections import defaultdict
from itertools import combinations

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as XLImage
from scipy.stats import friedmanchisquare, wilcoxon
import matplotlib
matplotlib.use("Agg")  # bez GUI/tkinter - staci na ulozenie PNG pre embed do xlsx
import matplotlib.pyplot as plt

from TSPCT_experiment_runner import style_header_row, style_body_cells, autofit_columns

RESULTS_XLSX = os.path.join("results", "TSPCT_results.xlsx")
CONVERGENCE_JSON = os.path.join("results", "convergence_curves.json")

ALGO_COLORS = {"ACO": "red", "GA": "green", "ABC": "magenta", "PSO": "blue",
               "FFA": "darkorange", "CSA": "cyan"}


# ----------------------------------------------------------
#  Nacitanie dat
# ----------------------------------------------------------
def load_raw_from_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Raw Results"]
    rows = []
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        d = dict(zip(headers, row))
        rows.append({
            "algorithm": d["Algorithm"], "environment": d["Environment"],
            "setting": d["Setting"], "run": int(d["Run"]), "seed": int(d["Seed"]),
            "final_distance": float(d["Final Distance"]),
            "best_gen": int(d["Best Generation"]), "time_s": float(d["Time (s)"]),
        })
    return rows


def load_convergence(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


# ----------------------------------------------------------
#  Friedman test + post-hoc Wilcoxon (Holm korekcia)
# ----------------------------------------------------------
def compute_statistical_tests(rows):
    groups = defaultdict(lambda: defaultdict(dict))
    for r in rows:
        groups[(r["environment"], r["setting"])][r["algorithm"]][r["run"]] = r["final_distance"]

    friedman_rows, wilcoxon_rows = [], []

    for (env, setting), algo_dict in sorted(groups.items()):
        algos = sorted(algo_dict.keys())
        run_id_sets = [set(algo_dict[a].keys()) for a in algos]
        common_runs = sorted(set.intersection(*run_id_sets)) if run_id_sets else []
        if len(common_runs) < 3 or len(algos) < 2:
            continue

        samples = [[algo_dict[a][run] for run in common_runs] for a in algos]
        stat, p = friedmanchisquare(*samples)
        friedman_rows.append({
            "environment": env, "setting": setting, "n_algorithms": len(algos),
            "n_runs": len(common_runs), "friedman_stat": round(stat, 4),
            "p_value": round(p, 6), "significant": p < 0.05,
        })

        if p < 0.05:
            pairs = list(combinations(algos, 2))
            raw = []
            for a, b in pairs:
                sa = [algo_dict[a][run] for run in common_runs]
                sb = [algo_dict[b][run] for run in common_runs]
                try:
                    wstat, wp = wilcoxon(sa, sb)
                except ValueError:
                    wstat, wp = float("nan"), 1.0
                raw.append((a, b, wstat, wp))

            order = sorted(range(len(raw)), key=lambda i: raw[i][3])
            m = len(raw)
            adj = [None] * m
            prev = 0.0
            for rank, idx in enumerate(order):
                adjusted = max(min(1.0, raw[idx][3] * (m - rank)), prev)
                adj[idx] = adjusted
                prev = adjusted

            for (a, b, wstat, wp), padj in zip(raw, adj):
                wilcoxon_rows.append({
                    "environment": env, "setting": setting, "algo_a": a, "algo_b": b,
                    "wilcoxon_stat": None if math.isnan(wstat) else round(wstat, 4),
                    "p_raw": round(wp, 6), "p_holm": round(padj, 6),
                    "significant": padj < 0.05,
                })

    return friedman_rows, wilcoxon_rows


# ----------------------------------------------------------
#  Zapis Friedman + Wilcoxon do zosita
# ----------------------------------------------------------
def write_test_sheets(wb, friedman_rows, wilcoxon_rows):
    for name in ("Friedman Test", "Wilcoxon Posthoc"):
        if name in wb.sheetnames:
            del wb[name]

    ws_f = wb.create_sheet("Friedman Test")
    f_headers = ["Environment", "Setting", "N Algorithms", "N Runs",
                 "Friedman Statistic", "p-value", "Significant (p<0.05)"]
    ws_f.append(f_headers)
    for r in friedman_rows:
        ws_f.append([r["environment"], r["setting"], r["n_algorithms"], r["n_runs"],
                     r["friedman_stat"], r["p_value"], "ANO" if r["significant"] else "nie"])
    style_header_row(ws_f)
    style_body_cells(ws_f)
    autofit_columns(ws_f)
    note_row = ws_f.max_row + 2
    ws_f.cell(row=note_row, column=1, value=(
        "Poznamka: Friedmanov test porovnava vsetky algoritmy naprieč parovanymi behmi "
        "(rovnaka instancia miest, rovnake cislo behu). Vypocitane cez "
        "scipy.stats.friedmanchisquare."
    )).font = Font(italic=True, size=9, color="666666")

    ws_w = wb.create_sheet("Wilcoxon Posthoc")
    w_headers = ["Environment", "Setting", "Algorithm A", "Algorithm B",
                 "Wilcoxon Statistic", "p (raw)", "p (Holm-corrected)", "Significant (p<0.05)"]
    ws_w.append(w_headers)
    for r in wilcoxon_rows:
        ws_w.append([r["environment"], r["setting"], r["algo_a"], r["algo_b"],
                     r["wilcoxon_stat"], r["p_raw"], r["p_holm"],
                     "ANO" if r["significant"] else "nie"])
    style_header_row(ws_w)
    style_body_cells(ws_w)
    autofit_columns(ws_w)
    note_row = ws_w.max_row + 2
    ws_w.cell(row=note_row, column=1, value=(
        "Poznamka: parove Wilcoxon testy sa robia iba pre kombinacie, kde vysiel Friedman "
        "test signifikantny (p<0.05). Holm-Bonferroni korekcia kontroluje familywise chybu."
    )).font = Font(italic=True, size=9, color="666666")


# ----------------------------------------------------------
#  Konvergencne krivky -> PNG -> vlozenie do xlsx
# ----------------------------------------------------------
def build_convergence_images(conv_data, tmp_dir):
    groups = defaultdict(lambda: defaultdict(list))
    for entry in conv_data:
        groups[(entry["environment"], entry["setting"])][entry["algorithm"]].append(entry["convergence"])

    image_paths = []
    for (env, setting), algo_curves in sorted(groups.items()):
        plt.figure(figsize=(7.5, 4.5))
        n = 0
        for algo, curves in sorted(algo_curves.items()):
            min_len = min(len(c) for c in curves)
            curves = [c[:min_len] for c in curves]
            n = len(curves)

            means, ci = [], []
            for gen_idx in range(min_len):
                vals = [c[gen_idx] for c in curves]
                m = sum(vals) / n
                se = (math.sqrt(sum((v - m) ** 2 for v in vals) / (n - 1)) / math.sqrt(n)) if n > 1 else 0.0
                means.append(m)
                ci.append(1.96 * se)

            xs = list(range(min_len))
            lower = [m - c for m, c in zip(means, ci)]
            upper = [m + c for m, c in zip(means, ci)]
            color = ALGO_COLORS.get(algo)
            plt.plot(xs, means, label=algo, color=color, linewidth=1.5)
            plt.fill_between(xs, lower, upper, color=color, alpha=0.15)

        plt.title(f"Priemerna konvergencia — {env}, {setting} (n={n} behov, 95% CI)")
        plt.xlabel("Generacia")
        plt.ylabel("Najlepsia dlzka trasy (priemer)")
        plt.legend()
        plt.grid(alpha=0.3)
        plt.tight_layout()
        out_path = os.path.join(tmp_dir, f"conv_{env}_{setting}.png")
        plt.savefig(out_path, dpi=140)
        plt.close()
        image_paths.append((env, setting, out_path))
    return image_paths


def write_convergence_sheet(wb, image_paths):
    if "Convergence Charts" in wb.sheetnames:
        del wb["Convergence Charts"]
    ws = wb.create_sheet("Convergence Charts")
    ws.column_dimensions["A"].width = 4
    row_cursor = 1
    for env, setting, path in image_paths:
        ws.cell(row=row_cursor, column=2, value=f"{env} — {setting}").font = Font(bold=True, size=12)
        row_cursor += 1
        img = XLImage(path)
        img.width, img.height = 560, 340
        ws.add_image(img, f"B{row_cursor}")
        row_cursor += 19  # cca vyska obrazku v riadkoch


# ----------------------------------------------------------
#  About / metodika hárok
# ----------------------------------------------------------
def write_about_sheet(wb):
    if "About" in wb.sheetnames:
        del wb["About"]
    ws = wb.create_sheet("About", 0)  # prvy harok v zosite
    ws.column_dimensions["A"].width = 100
    lines = [
        ("TSPCT — vysledky vedeckeho benchmarku", True, 14),
        ("", False, 11),
        ("Metodika:", True, 12),
        ("- 3 prostredia (50 / 100 / 150 miest), suradnice fixne pre dane prostredie.", False, 11),
        ("- 3 nastavenia hyperparametrov (basic / fast_convergence / diversity).", False, 11),
        ("- 20 nezavislych behov na kombinaciu (algoritmus x prostredie x nastavenie),", False, 11),
        ("  lisiacich sa iba internym random seedom algoritmu (1..20).", False, 11),
        ("- Kazdy algoritmus rieši v danom behu presne tu istu instanciu problemu ako", False, 11),
        ("  ostatne algoritmy (parovane porovnanie).", False, 11),
        ("", False, 11),
        ("Harky v tomto zosite:", True, 12),
        ("- Raw Results: 1 riadok = 1 beh (algoritmus, prostredie, nastavenie, run, seed,", False, 11),
        ("  vysledna dlzka trasy, generacia najlepsieho riesenia, cas behu).", False, 11),
        ("- Aggregate Stats: mean/std/min/max/pocet behov na kombinaciu — pocitane", False, 11),
        ("  Excel formulami (AVERAGEIFS/SUMPRODUCT/MINIFS/MAXIFS) priamo z Raw Results,", False, 11),
        ("  takze sa prepocitaju automaticky pri zmene dat.", False, 11),
        ("- Friedman Test: test signifikancie rozdielov medzi algoritmami (scipy.stats).", False, 11),
        ("- Wilcoxon Posthoc: parove porovnania algoritmov s Holm korekciou (iba tam,", False, 11),
        ("  kde Friedman test vysiel signifikantny).", False, 11),
        ("- Convergence Charts: priemerna konvergencna krivka + 95% interval spolahlivosti", False, 11),
        ("  pre kazdu kombinaciu prostredie x nastavenie.", False, 11),
    ]
    for i, (text, bold, size) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(bold=bold, size=size)
        c.alignment = Alignment(wrap_text=False)


# ----------------------------------------------------------
#  Main
# ----------------------------------------------------------
def main():
    if not os.path.exists(RESULTS_XLSX):
        raise SystemExit(f"Nenajdeny {RESULTS_XLSX} - najprv spusti TSPCT_experiment_runner.py")

    rows = load_raw_from_xlsx(RESULTS_XLSX)
    conv = load_convergence(CONVERGENCE_JSON)

    friedman_rows, wilcoxon_rows = compute_statistical_tests(rows)

    wb = openpyxl.load_workbook(RESULTS_XLSX)
    write_test_sheets(wb, friedman_rows, wilcoxon_rows)

    with tempfile.TemporaryDirectory() as tmp_dir:
        image_paths = build_convergence_images(conv, tmp_dir)
        write_convergence_sheet(wb, image_paths)
        write_about_sheet(wb)
        wb.save(RESULTS_XLSX)

    print("Hotovo. Doplnene harky (Friedman Test, Wilcoxon Posthoc, Convergence Charts, About)")
    print(f"do: {RESULTS_XLSX}")


if __name__ == "__main__":
    main()