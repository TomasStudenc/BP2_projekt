# ==========================================================
# TSPCT_experiment_runner.py
#
#  Systematicke opakovane experimenty pre porovnanie
#  algoritmov ACO, GA, ABC, PSO na probleme obchodneho
#  cestujuceho.
#
#  Navrh:
#   - 3 prostredia (50 / 100 / 150 miest). Suradnice miest su
#     FIXNE pre dane prostredie (rovnake pre vsetky algoritmy
#     aj vsetky behy) - zabezpecuje to korektne porovnanie
#     a umoznuje parove statisticke testy.
#   - 3 nastavenia hyperparametrov na prostredie
#     (basic / fast_convergence / diversity)
#   - 15 nezavislych behov na kombinaciu (algoritmus x
#     prostredie x nastavenie). Behy sa lisia iba vnutornym
#     random seedom algoritmu (10001, 10002, ... 10015).
#   - Vysledky: raw_results.csv (jeden riadok = jeden beh),
#     aggregate_stats.csv (priemer/std/min/max/median),
#     convergence_curves.json (cele konvergencne krivky
#     vsetkych behov, na neskorsie priemerovanie/vykreslenie).
#
#  POUZITIE:
#   1. Uprav PARAM_SETS nizsie na presne hodnoty, ktore chces
#      pouzit pre "basic" / "fast_convergence" / "diversity".
#   2. (Volitelne) uprav RUNS_PER_CONFIG, ENVIRONMENTS, WORKERS.
#   3. Spusti:  python TSPCT_experiment_runner.py
#      Vysledky sa ulozia do priecinka ./results
# ==========================================================

import random
import time
import json
import os
import statistics
from collections import defaultdict
from concurrent.futures import ProcessPoolExecutor, as_completed

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import TSPCT_ACO
import TSPCT_GA
import TSPCT_ABC
import TSPCT_PSO
# Ak chces zahrnut aj FFA a CSA do experimentov, odkomentuj:
# import TSPCT_FFA
# import TSPCT_CSA

# ----------------------------------------------------------
#  KONFIGURACIA EXPERIMENTU
# ----------------------------------------------------------

RUNS_PER_CONFIG = 20       # pocet nezavislych behov na kombinaciu
ENV_SIZE = 500             # velkost kresliacej plochy (ako v aplikacii)
WORKERS = None             # None = pouzije vsetky dostupne CPU jadra
SEED_OFFSET = 0            # ACO_seed/GA_seed/... pojdu 1..RUNS_PER_CONFIG (+offset)

# algoritmus -> (funkcia, nazov klaca pre seed v params dict)
ALGORITHMS = {
    "ACO": (TSPCT_ACO.ACO, "ACO_seed"),
    "GA":  (TSPCT_GA.GA,   "GA_seed"),
    "ABC": (TSPCT_ABC.ABC, "ABC_seed"),
    "PSO": (TSPCT_PSO.PSO, "PSO_seed"),
    # "FFA": (TSPCT_FFA.FFA, "FFA_seed"),
    # "CSA": (TSPCT_CSA.CSA, "CSA_seed"),
}

# 3 prostredia - kazde ma vlastny FIXNY seed pre generovanie
# suradnic miest (rovnaka instancia pre vsetky algoritmy/behy)
ENVIRONMENTS = {
    "env1_50":  {"n_cities": 50,  "instance_seed": 1001},
    "env2_100": {"n_cities": 100, "instance_seed": 1002},
    "env3_150": {"n_cities": 150, "instance_seed": 1003},
}

# ----------------------------------------------------------
#  !! DOPLN PRESNE HODNOTY HYPERPARAMETROV !!
#  Nizsie su iba orientacne (placeholder) hodnoty v duchu
#  povodnej clankovej logiky:
#   - basic            = povodne defaultne hodnoty z GUI
#   - fast_convergence = vacsia populacia / silnejsi tlak
#                        na najlepsie riesenia
#   - diversity        = vyssia mutacia / nahodnost, mensi
#                         selekcny tlak
#  Uprav podla toho, co chces v clanku prezentovat (mozes
#  pouzit presne tie iste hodnoty ako v povodnom experimente,
#  ak si ich mas zapisane).
# ----------------------------------------------------------

PARAM_SETS = {
    "basic": {
        "ACO": {"ACO generations": 1000, "Ant count": 5, "Alpha": 1.0,
                 "Beta": 2.0, "Evaporation rate": 0.1, "Q": 100.0},
        "GA":  {"GA generations": 1000, "Population size": 100,
                 "Elit rate": 0.25, "Mutation rate": 0.1},
        "ABC": {"ABC generations": 1000, "Bee count": 20,
                 "employ_rate": 0.7, "scout_rate": 0.01},
        "PSO": {"PSO generations": 1000, "Particle count": 10,
                 "c1": 1.5, "c2": 1.5, "weight": 0.9},
    },
    "fast_convergence": {
        "ACO": {"ACO generations": 1000, "Ant count": 15, "Alpha": 1.5,
                 "Beta": 3.0, "Evaporation rate": 0.05, "Q": 150.0},
        "GA":  {"GA generations": 1000, "Population size": 100,
                 "Elit rate": 0.45, "Mutation rate": 0.03},
        "ABC": {"ABC generations": 1000, "Bee count": 20,
                 "employ_rate": 0.85, "scout_rate": 0.005},
        "PSO": {"PSO generations": 1000, "Particle count": 15,
                 "c1": 1.0, "c2": 2.5, "weight": 0.6},
    },
    "diversity": {
        "ACO": {"ACO generations": 1000, "Ant count": 25, "Alpha": 0.5,
                 "Beta": 1.5, "Evaporation rate": 0.35, "Q": 70.0},
        "GA":  {"GA generations": 1000, "Population size": 200,
                 "Elit rate": 0.1, "Mutation rate": 0.35},
        "ABC": {"ABC generations": 1000, "Bee count": 50,
                 "employ_rate": 0.5, "scout_rate": 0.08},
        "PSO": {"PSO generations": 1000, "Particle count": 30,
                 "c1": 2.2, "c2": 0.8, "weight": 0.95},
    },
}


# ----------------------------------------------------------
#  Generovanie fixnej instancie miest pre prostredie
# ----------------------------------------------------------
def generate_instance(n_cities, size, seed):
    rnd = random.Random(seed)
    cords = []
    while len(cords) < n_cities:
        p = (rnd.randint(0, size), rnd.randint(0, size))
        if p not in cords:
            cords.append(p)
    return cords


# ----------------------------------------------------------
#  Jeden nezavisly beh (spusteny v samostatnom procese)
# ----------------------------------------------------------
def run_single(job):
    algo_name, env_name, setting_name, run_idx, cords, base_params, seed_key = job
    func, _ = ALGORITHMS[algo_name]

    params = dict(base_params)
    params[seed_key] = SEED_OFFSET + run_idx  # ACO_seed/GA_seed/... = 1..RUNS_PER_CONFIG

    start = time.time()
    route, distances, best_gen = func(cords, params)
    elapsed = time.time() - start

    return {
        "algorithm": algo_name,
        "environment": env_name,
        "setting": setting_name,
        "run": run_idx,
        "seed": params[seed_key],
        "final_distance": distances[-1],
        "best_gen": best_gen,
        "time_s": round(elapsed, 3),
        "convergence": distances,
    }


# ----------------------------------------------------------
#  Hlavny beh experimentu
# ----------------------------------------------------------
def build_jobs():
    jobs = []
    instances = {}
    for env_name, env_cfg in ENVIRONMENTS.items():
        cords = generate_instance(env_cfg["n_cities"], ENV_SIZE, env_cfg["instance_seed"])
        instances[env_name] = cords
        for setting_name, algo_params in PARAM_SETS.items():
            for algo_name in ALGORITHMS:
                base_params = algo_params[algo_name]
                for run_idx in range(1, RUNS_PER_CONFIG + 1):
                    jobs.append((algo_name, env_name, setting_name, run_idx,
                                 cords, base_params, ALGORITHMS[algo_name][1]))
    return jobs, instances


def main(save_convergence=True, workers=WORKERS):
    jobs, instances = build_jobs()
    print(f"Naplanovanych behov spolu: {len(jobs)}")

    results = []
    with ProcessPoolExecutor(max_workers=workers) as ex:
        futures = {ex.submit(run_single, job): job for job in jobs}
        done = 0
        for fut in as_completed(futures):
            res = fut.result()
            results.append(res)
            done += 1
            if done % 10 == 0 or done == len(jobs):
                print(f"[{done}/{len(jobs)}] posledny: {res['algorithm']} "
                      f"{res['environment']} {res['setting']} run {res['run']} "
                      f"-> dist={res['final_distance']:.2f}, t={res['time_s']}s")

    os.makedirs("results", exist_ok=True)

    xlsx_path = os.path.join("results", "TSPCT_results.xlsx")
    write_results_workbook(results, xlsx_path)
    print(f"Ulozene: {xlsx_path}  (hárky: Raw Results, Aggregate Stats)")

    if save_convergence:
        conv_path = os.path.join("results", "convergence_curves.json")
        with open(conv_path, "w", encoding="utf-8") as f:
            json.dump(
                [{"algorithm": r["algorithm"], "environment": r["environment"],
                  "setting": r["setting"], "run": r["run"],
                  "convergence": r["convergence"]} for r in results],
                f
            )
        print(f"Ulozene: {conv_path}  (pouzije TSPCT_analyze_results.py)")

    return results


# ----------------------------------------------------------
#  Styling helpers (spolocne pouzivane aj v analyze skripte)
# ----------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Calibri", size=10)
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row=1, ncols=None):
    ncols = ncols or ws.max_column
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws.freeze_panes = ws.cell(row=row + 1, column=1).coordinate
    ws.auto_filter.ref = ws.dimensions


def autofit_columns(ws, min_width=10, max_width=32):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            col = cell.column_letter
            widths[col] = max(widths.get(col, min_width), min(length + 3, max_width))
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def style_body_cells(ws, row_start=2, row_end=None, col_end=None):
    row_end = row_end or ws.max_row
    col_end = col_end or ws.max_column
    for row in ws.iter_rows(min_row=row_start, max_row=row_end, max_col=col_end):
        for cell in row:
            cell.font = BODY_FONT
            cell.border = BORDER


# ----------------------------------------------------------
#  Zapis vysledkov do Excelu (Raw Results + Aggregate Stats)
# ----------------------------------------------------------
def write_results_workbook(results, path):
    wb = openpyxl.Workbook()

    # -------- Raw Results --------
    ws_raw = wb.active
    ws_raw.title = "Raw Results"
    headers = ["Algorithm", "Environment", "Setting", "Run", "Seed",
               "Final Distance", "Best Generation", "Time (s)"]
    ws_raw.append(headers)
    for r in results:
        ws_raw.append([r["algorithm"], r["environment"], r["setting"], r["run"],
                        r["seed"], r["final_distance"], r["best_gen"], r["time_s"]])
    n_raw = len(results)
    for row in range(2, n_raw + 2):
        ws_raw.cell(row=row, column=6).number_format = "0.00"
        ws_raw.cell(row=row, column=8).number_format = "0.000"
    style_header_row(ws_raw)
    style_body_cells(ws_raw)
    autofit_columns(ws_raw)

    # -------- Aggregate Stats (formulas referencing Raw Results) --------
    ws_agg = wb.create_sheet("Aggregate Stats")
    agg_headers = ["Algorithm", "Environment", "Setting", "N",
                   "Distance Mean", "Distance Std", "Distance Min", "Distance Max",
                   "Time Mean (s)"]
    ws_agg.append(agg_headers)

    combos = sorted({(r["algorithm"], r["environment"], r["setting"]) for r in results})
    last_raw_row = n_raw + 1  # +1 for header

    A, E, S, D, T = "A", "B", "C", "F", "H"  # stlpce v "Raw Results": Algo,Env,Setting,...,Distance(F),...,Time(H)
    for i, (algo, env, setting) in enumerate(combos, start=2):
        ws_agg.cell(row=i, column=1, value=algo)
        ws_agg.cell(row=i, column=2, value=env)
        ws_agg.cell(row=i, column=3, value=setting)

        rng = f"'Raw Results'!${A}$2:${A}${last_raw_row}"
        rng_env = f"'Raw Results'!${E}$2:${E}${last_raw_row}"
        rng_set = f"'Raw Results'!${S}$2:${S}${last_raw_row}"
        rng_dist = f"'Raw Results'!${D}$2:${D}${last_raw_row}"
        rng_time = f"'Raw Results'!${T}$2:${T}${last_raw_row}"

        a_cell, e_cell, s_cell = f"$A{i}", f"$B{i}", f"$C{i}"

        n_formula = f"=COUNTIFS({rng},{a_cell},{rng_env},{e_cell},{rng_set},{s_cell})"
        mean_formula = (f"=AVERAGEIFS({rng_dist},{rng},{a_cell},{rng_env},{e_cell},"
                         f"{rng_set},{s_cell})")
        # Std (vzorkova smerodajna odchylka) pomocou SUMPRODUCT - Excel 2007-era,
        # funguje bez CSE a bez XLOOKUP/array-spill funkcii.
        std_formula = (
            f"=SQRT(SUMPRODUCT(({rng_env}={e_cell})*({rng_set}={s_cell})*"
            f"({rng}={a_cell})*({rng_dist}-E{i})^2)/(D{i}-1))"
        )
        min_formula = f"=_xlfn.MINIFS({rng_dist},{rng},{a_cell},{rng_env},{e_cell},{rng_set},{s_cell})"
        max_formula = f"=_xlfn.MAXIFS({rng_dist},{rng},{a_cell},{rng_env},{e_cell},{rng_set},{s_cell})"
        time_formula = (f"=AVERAGEIFS({rng_time},{rng},{a_cell},{rng_env},{e_cell},"
                         f"{rng_set},{s_cell})")

        ws_agg.cell(row=i, column=4, value=n_formula)
        ws_agg.cell(row=i, column=5, value=mean_formula).number_format = "0.00"
        ws_agg.cell(row=i, column=6, value=std_formula).number_format = "0.00"
        ws_agg.cell(row=i, column=7, value=min_formula).number_format = "0.00"
        ws_agg.cell(row=i, column=8, value=max_formula).number_format = "0.00"
        ws_agg.cell(row=i, column=9, value=time_formula).number_format = "0.000"

    style_header_row(ws_agg)
    style_body_cells(ws_agg)
    autofit_columns(ws_agg)

    wb.save(path)


if __name__ == "__main__":
    main()